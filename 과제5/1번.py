from openpyxl import load_workbook
import matplotlib.pyplot as plt

while True:
    load_wb = load_workbook("elec.xlsx", data_only=True)
    load_ws = load_wb['Sheet1']
    labels=load_ws.cell(1, 1).value, load_ws.cell(1, 2).value, load_ws.cell(1, 3).value, load_ws.cell(1, 4).value, load_ws.cell(1, 5).value
    num=[load_ws.cell(2, 1).value,load_ws.cell(2, 2).value,load_ws.cell(2, 3).value,load_ws.cell(2, 4).value,load_ws.cell(2, 5).value]
    explode = (0.1, 0.1, 0.1, 0.1, 0.1)
    plt.pie(num, explode=explode, labels=labels, autopct='%1.1f%%', startangle=67) 
    plt.draw()              # 그리기
    plt.pause(10)         # 잠시  기다리기
    plt.clf()                # figure 지우기

